import json
import base64
import os
import calendar
from datetime import date, datetime
from decimal import Decimal
from urllib.parse import quote

import requests as http_requests
from django.conf import settings
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

from .auth import google_auth
from .models import User, Expense, Budget, PushSubscription
from .groq_helper import groq_chat, parse_json, CATEGORIES, EXTRACT_PROMPT


def _json(request):
    return json.loads(request.body) if request.body else {}


def _expense_dict(e):
    return {
        'id': e.id, 'user_id': e.user_id, 'description': e.description,
        'amount': float(e.amount), 'category': e.category,
        'date': e.date.isoformat() if hasattr(e.date, 'isoformat') else str(e.date) if e.date else None,
        'product_name': e.product_name, 'quantity': float(e.quantity),
        'size': e.size, 'mrp': float(e.mrp) if e.mrp else float(e.amount),
    }


import hashlib
import uuid

# ─── Auth ───
@csrf_exempt
@google_auth
def auth_view(request):
    u = request.guser
    return JsonResponse({'google_id': u.google_id, 'email': u.email, 'name': u.name, 'picture': u.picture})


def _hash_pw(password):
    return hashlib.sha256(password.encode()).hexdigest()


@csrf_exempt
def register_view(request):
    d = _json(request)
    email, password, name = d.get('email', '').strip().lower(), d.get('password', ''), d.get('name', '').strip()
    if not email or not password or not name:
        return JsonResponse({'error': 'Email, password and name required'}, status=400)
    if len(password) < 6:
        return JsonResponse({'error': 'Password must be at least 6 characters'}, status=400)
    if User.objects.filter(email=email, password_hash__gt='').exists():
        return JsonResponse({'error': 'Email already registered'}, status=409)
    uid = f'email_{uuid.uuid4().hex[:16]}'
    user = User.objects.create(google_id=uid, email=email, name=name, password_hash=_hash_pw(password))
    return JsonResponse({'google_id': user.google_id, 'email': user.email, 'name': user.name, 'picture': '', 'token': uid})


@csrf_exempt
def login_view(request):
    d = _json(request)
    email, password = d.get('email', '').strip().lower(), d.get('password', '')
    if not email or not password:
        return JsonResponse({'error': 'Email and password required'}, status=400)
    try:
        user = User.objects.get(email=email, password_hash=_hash_pw(password))
        return JsonResponse({'google_id': user.google_id, 'email': user.email, 'name': user.name, 'picture': user.picture, 'token': user.google_id})
    except User.DoesNotExist:
        return JsonResponse({'error': 'Invalid email or password'}, status=401)


# ─── Expenses CRUD ───
@csrf_exempt
@google_auth
def expenses_list_or_create(request):
    if request.method == 'GET':
        month = request.GET.get('month')
        qs = Expense.objects.filter(user_id=request.guser.google_id)
        if month:
            qs = qs.filter(date__year=int(month[:4]), date__month=int(month[5:7]))
        return JsonResponse([_expense_dict(e) for e in qs], safe=False)
    d = _json(request)
    e = Expense.objects.create(
        user_id=request.guser.google_id, description=d.get('description', ''),
        amount=d.get('amount', 0), category=d.get('category', 'Other'),
        date=d.get('date', date.today().isoformat()),
        product_name=d.get('product_name') or d.get('description', ''),
        quantity=d.get('quantity', 1), size=d.get('size'),
        mrp=d.get('mrp') or d.get('amount', 0),
    )
    return JsonResponse(_expense_dict(e))


@csrf_exempt
@google_auth
def expenses_update_or_delete(request, pk):
    if request.method == 'DELETE':
        Expense.objects.filter(id=pk, user_id=request.guser.google_id).delete()
        return JsonResponse({'ok': True})
    d = _json(request)
    qs = Expense.objects.filter(id=pk, user_id=request.guser.google_id)
    if not qs.exists():
        return JsonResponse({'error': 'Not found'}, status=404)
    qs.update(
        description=d.get('description', ''), amount=d.get('amount', 0),
        category=d.get('category', 'Other'), date=d.get('date'),
        product_name=d.get('product_name') or d.get('description', ''),
        quantity=d.get('quantity', 1), size=d.get('size'),
        mrp=d.get('mrp') or d.get('amount', 0),
    )
    return JsonResponse(_expense_dict(qs.first()))


@csrf_exempt
@google_auth
def expenses_bulk(request):
    items = _json(request).get('items', [])
    if not items:
        return JsonResponse({'error': 'No items'}, status=400)
    results = []
    for d in items:
        e = Expense.objects.create(
            user_id=request.guser.google_id, description=d.get('description', ''),
            amount=d.get('amount', 0), category=d.get('category', 'Other'),
            date=d.get('date', date.today().isoformat()),
            product_name=d.get('product_name') or d.get('description', ''),
            quantity=d.get('quantity', 1), size=d.get('size'),
            mrp=d.get('mrp') or d.get('amount', 0),
        )
        results.append(_expense_dict(e))
    return JsonResponse(results, safe=False)


# ─── Budget ───
@csrf_exempt
@google_auth
def budget_get_or_set(request):
    if request.method == 'GET':
        month = request.GET.get('month', date.today().strftime('%Y-%m'))
        try:
            b = Budget.objects.get(user_id=request.guser.google_id, month=month)
            return JsonResponse({'salary': float(b.salary), 'budget': float(b.budget), 'month': b.month})
        except Budget.DoesNotExist:
            return JsonResponse({'salary': 0, 'budget': 0, 'month': month})
    d = _json(request)
    b, _ = Budget.objects.update_or_create(
        user_id=request.guser.google_id, month=d['month'],
        defaults={'salary': d.get('salary', 0), 'budget': d.get('budget', 0)},
    )
    return JsonResponse({'salary': float(b.salary), 'budget': float(b.budget), 'month': b.month})


# ─── Dashboard ───
MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


@csrf_exempt
@google_auth
def dashboard(request):
    month = request.GET.get('month', date.today().strftime('%Y-%m'))
    uid = request.guser.google_id
    year, mon = int(month[:4]), int(month[5:7])

    try:
        b = Budget.objects.get(user_id=uid, month=month)
        budget_salary, budget_amount = float(b.salary), float(b.budget)
    except Budget.DoesNotExist:
        budget_salary, budget_amount = 0, 0

    expenses = list(Expense.objects.filter(user_id=uid, date__year=year, date__month=mon).order_by('date'))
    exp_dicts = [_expense_dict(e) for e in expenses]

    total_spent = sum(e['amount'] for e in exp_dicts)
    total_mrp = sum(e['mrp'] for e in exp_dicts)
    total_discount = total_mrp - total_spent
    remaining = budget_amount - total_spent
    utilization = (total_spent / budget_amount * 100) if budget_amount > 0 else 0

    # Category breakdown
    cat_map = {}
    for e in exp_dicts:
        cat_map[e['category']] = cat_map.get(e['category'], 0) + e['amount']
    categories = sorted(
        [{'category': c, 'total': round(t, 2), 'percent': round(t / total_spent * 100, 2) if total_spent > 0 else 0} for c, t in cat_map.items()],
        key=lambda x: -x['total']
    )

    # Top 5
    top5 = sorted(exp_dicts, key=lambda e: -e['amount'])[:5]
    top5 = [{'product_name': e['product_name'] or e['description'], 'category': e['category'], 'amount': e['amount']} for e in top5]

    # Weekly forecast
    days_in_month = calendar.monthrange(year, mon)[1]
    today = date.today()
    weeks = []
    week_start = 1
    week_num = 1
    while week_start <= days_in_month:
        week_end = min(week_start + 6, days_in_month)
        start_date = f"{month}-{week_start:02d}"
        end_date = f"{month}-{week_end:02d}"
        week_exps = [e for e in exp_dicts if week_start <= int(e['date'][8:10]) <= week_end]
        actual = sum(e['amount'] for e in week_exps)

        past_exps = [e for e in exp_dicts if e['date'] < start_date]
        days_so_far = max(week_start - 1, 1)
        past_total = sum(e['amount'] for e in past_exps)
        avg_daily = past_total / days_so_far if past_total > 0 else (budget_amount / days_in_month if budget_amount else 0)
        forecast = round(avg_daily * (week_end - week_start + 1))

        week_end_date = date(year, mon, week_end)
        status = 'covered' if week_end_date < today else ('in_progress' if date(year, mon, week_start) <= today else 'upcoming')

        weeks.append({
            'week': week_num, 'startDate': start_date, 'endDate': end_date,
            'dateRange': f"{MONTH_NAMES[mon - 1]} {week_start} – {MONTH_NAMES[mon - 1]} {week_end}",
            'forecast': forecast, 'actual': round(actual), 'variance': round(actual - forecast), 'status': status,
        })
        week_start = week_end + 1
        week_num += 1

    # Next week prediction
    completed = [w for w in weeks if w['status'] == 'covered' and w['actual'] > 0]
    next_week = next((w for w in weeks if w['status'] == 'upcoming'), None) or next((w for w in weeks if w['status'] == 'in_progress'), None)
    prediction = None
    if completed and next_week:
        weighted_sum = sum(w['actual'] * (i + 1) for i, w in enumerate(completed))
        weight_total = sum(i + 1 for i in range(len(completed)))
        predicted = round(weighted_sum / weight_total)
        trend = completed[-1]['actual'] - completed[-2]['actual'] if len(completed) >= 2 else 0
        remaining_weeks = len([w for w in weeks if w['status'] != 'covered'])
        eom = round(total_spent + predicted * remaining_weeks)
        prediction = {
            'week': next_week['week'], 'dateRange': next_week['dateRange'], 'predicted': predicted,
            'trend': 'up' if trend > 0 else ('down' if trend < 0 else 'stable'),
            'trendAmount': abs(round(trend)), 'endOfMonthProjection': eom,
            'willExceedBudget': budget_amount > 0 and eom > budget_amount,
            'overBy': max(0, round(eom - budget_amount)),
            'topCategory': categories[0]['category'] if categories else 'N/A',
            'basedOnWeeks': len(completed),
        }

    return JsonResponse({
        'budget': {'salary': budget_salary, 'budget': budget_amount},
        'totalSpent': round(total_spent, 2), 'totalMRP': round(total_mrp, 2),
        'totalDiscount': round(total_discount, 2), 'remaining': round(remaining, 2),
        'utilization': round(utilization, 2), 'categories': categories, 'top5': top5,
        'weeks': weeks, 'nextWeekPrediction': prediction, 'expenseCount': len(expenses),
    })


# ─── AI Insights ───
@csrf_exempt
@google_auth
def insights(request):
    month = request.GET.get('month', date.today().strftime('%Y-%m'))
    uid = request.guser.google_id
    year, mon = int(month[:4]), int(month[5:7])

    try:
        b = Budget.objects.get(user_id=uid, month=month)
        budget_amount, salary = float(b.budget), float(b.salary)
    except Budget.DoesNotExist:
        budget_amount, salary = 0, 0

    from django.db.models import Sum
    cat_totals = Expense.objects.filter(user_id=uid, date__year=year, date__month=mon).values('category').annotate(total=Sum('amount')).order_by('-total')
    total_spent = sum(float(c['total']) for c in cat_totals)
    breakdown = ', '.join(f"{c['category']}: ₹{float(c['total']):.0f}" for c in cat_totals)

    try:
        raw = groq_chat([
            {'role': 'system', 'content': 'You are a concise Indian personal finance advisor. Given monthly spending data, return a JSON array of 4-6 insights. Each: {"type":"spending"|"alert"|"saving"|"forecast","icon":"📊"|"⚠️"|"💡"|"📈","title":"short title","text":"1-2 sentence insight"}. Return ONLY JSON array.'},
            {'role': 'user', 'content': f"Monthly budget: ₹{budget_amount}, Salary: ₹{salary}, Total spent: ₹{total_spent:.0f}, Remaining: ₹{budget_amount - total_spent:.0f}, Utilization: {(total_spent / budget_amount * 100):.1f}%\nBreakdown: {breakdown}" if budget_amount else f"Total spent: ₹{total_spent:.0f}\nBreakdown: {breakdown}"},
        ])
        return JsonResponse({'insights': parse_json(raw) or []})
    except Exception:
        return JsonResponse({'insights': [{'type': 'spending', 'icon': '📊', 'title': 'Budget Status', 'text': f"You have spent ₹{total_spent:.0f} of ₹{budget_amount} budget."}]})


# ─── Smart Categorization ───
@csrf_exempt
@google_auth
def categorize(request):
    d = _json(request)
    product_name = d.get('product_name', '')
    if not product_name:
        return JsonResponse({'category': 'Other'})
    try:
        raw = groq_chat([
            {'role': 'system', 'content': f"Categorize this Indian product into exactly one category. Categories: {', '.join(CATEGORIES)}. Return ONLY the category name, nothing else."},
            {'role': 'user', 'content': product_name},
        ], max_tokens=20)
        cat = raw.strip()
        return JsonResponse({'category': cat if cat in CATEGORIES else 'Other'})
    except Exception:
        return JsonResponse({'category': 'Other'})


# ─── Upload ───
@csrf_exempt
@google_auth
def upload(request):
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file'}, status=400)

    path = os.path.join(settings.MEDIA_ROOT, f.name)
    with open(path, 'wb') as dest:
        for chunk in f.chunks():
            dest.write(chunk)

    try:
        items = []
        mime = f.content_type or ''

        if mime.startswith('image/'):
            with open(path, 'rb') as fh:
                b64 = base64.b64encode(fh.read()).decode()
            raw = groq_chat([{'role': 'user', 'content': [
                {'type': 'text', 'text': EXTRACT_PROMPT},
                {'type': 'image_url', 'image_url': {'url': f'data:{mime};base64,{b64}'}},
            ]}], max_tokens=2000)
            items = parse_json(raw) or []

        elif mime == 'application/pdf' or f.name.lower().endswith('.pdf'):
            from PyPDF2 import PdfReader
            reader = PdfReader(path)
            text = '\n'.join(page.extract_text() or '' for page in reader.pages)
            raw = groq_chat([{'role': 'user', 'content': EXTRACT_PROMPT + '\n\nReceipt text:\n' + text[:4000]}], max_tokens=2000)
            items = parse_json(raw) or []

        elif f.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            headers = [str(c.value or '').lower() for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rd = dict(zip(headers, row))
                desc = rd.get('description') or rd.get('item') or rd.get('product name') or (row[0] if row else '')
                amt = rd.get('amount') or rd.get('price') or rd.get('total') or (row[1] if len(row) > 1 else 0)
                if not desc or not amt:
                    continue
                items.append({'description': str(desc), 'amount': float(str(amt).replace('₹', '').replace(',', '') or 0), 'category': rd.get('category', 'Other'), 'quantity': rd.get('quantity', 1) or 1, 'size': rd.get('size', ''), 'mrp': rd.get('mrp') or amt})

        today_str = date.today().isoformat()
        items = [
            {**i, 'date': today_str, 'category': i.get('category') if i.get('category') in CATEGORIES else 'Other', 'mrp': i.get('mrp') or i.get('amount', 0), 'quantity': i.get('quantity', 1)}
            for i in items if i.get('description') and (i.get('amount') or 0) > 0
        ]
        return JsonResponse({'items': items})
    except Exception as e:
        return JsonResponse({'error': f'Failed to parse: {e}'}, status=500)
    finally:
        os.unlink(path) if os.path.exists(path) else None


# ─── Email Sync (Swiggy, Zepto) ───
SYNC_PROVIDERS = {
    'swiggy': {'from': 'noreply@swiggy.in', 'label': 'Swiggy'},
    'zepto': {'from': 'noreply@zeptonow.com', 'label': 'Zepto'},
}


def _gmail_fetch(access_token, url):
    resp = http_requests.get(url, headers={'Authorization': f'Bearer {access_token}'}, timeout=30)
    data = resp.json()
    if not resp.ok:
        raise Exception(data.get('error', {}).get('message', f'Gmail API {resp.status_code}'))
    return data


def _find_pdf_attachment(payload):
    for part in payload.get('parts', []):
        if part.get('filename', '').lower().endswith('.pdf') and part.get('body', {}).get('attachmentId'):
            return {'id': part['body']['attachmentId'], 'name': part['filename']}
        found = _find_pdf_attachment(part)
        if found:
            return found
    return None


def _parse_pdf_bytes(pdf_bytes):
    import io
    from PyPDF2 import PdfReader
    reader = PdfReader(io.BytesIO(pdf_bytes))
    return '\n'.join(page.extract_text() or '' for page in reader.pages)


@csrf_exempt
@google_auth
def email_sync(request, provider_key):
    provider = SYNC_PROVIDERS.get(provider_key)
    if not provider:
        return JsonResponse({'error': 'Unknown provider'}, status=400)

    d = _json(request)
    gmail_token = d.get('gmail_token')
    if not gmail_token:
        return JsonResponse({'error': 'Gmail token required'}, status=400)

    month = d.get('month', date.today().strftime('%Y-%m'))
    year, mon = int(month[:4]), int(month[5:7])
    last_day = calendar.monthrange(year, mon)[1]
    after = f"{year}/{mon:02d}/01"
    before = f"{year}/{mon:02d}/{last_day}"

    try:
        query = f"from:{provider['from']} has:attachment filename:pdf after:{after} before:{before}"
        list_data = _gmail_fetch(gmail_token, f"https://gmail.googleapis.com/gmail/v1/users/me/messages?q={quote(query)}&maxResults=50")

        if not list_data.get('messages'):
            return JsonResponse({'items': [], 'message': f"No {provider['label']} invoices found"})

        items = []
        for msg in list_data['messages']:
            try:
                detail = _gmail_fetch(gmail_token, f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg['id']}?format=full")
                headers = {h['name'].lower(): h['value'] for h in detail.get('payload', {}).get('headers', [])}
                email_date = datetime.strptime(headers.get('date', ''), '%a, %d %b %Y %H:%M:%S %z').strftime('%Y-%m-%d') if headers.get('date') else f"{year}-{mon:02d}-01"

                pdf = _find_pdf_attachment(detail.get('payload', {}))
                if not pdf:
                    continue

                att_data = _gmail_fetch(gmail_token, f"https://gmail.googleapis.com/gmail/v1/users/me/messages/{msg['id']}/attachments/{pdf['id']}")
                pdf_bytes = base64.urlsafe_b64decode(att_data['data'] + '==')
                text = _parse_pdf_bytes(pdf_bytes)

                raw = groq_chat([{'role': 'user', 'content': EXTRACT_PROMPT + '\n\nReceipt text:\n' + text[:4000]}], max_tokens=2000)
                parsed = parse_json(raw) or []

                for item in parsed:
                    if not item.get('description') or not item.get('amount') or item['amount'] <= 0:
                        continue
                    items.append({
                        **item, 'date': email_date,
                        'category': item.get('category') if item.get('category') in CATEGORIES else 'Other',
                        'mrp': item.get('mrp') or item['amount'], 'quantity': item.get('quantity', 1),
                        'email_id': msg['id'],
                    })
            except Exception:
                continue

        return JsonResponse({'items': items, 'total': len(items)})
    except Exception as e:
        return JsonResponse({'error': f'Failed to fetch emails: {e}'}, status=500)


@csrf_exempt
@google_auth
def sync_status(request):
    if request.method == 'GET':
        return JsonResponse({'synced': request.guser.synced_providers or []})
    d = _json(request)
    provider = d.get('provider')
    user = request.guser
    if provider and provider not in (user.synced_providers or []):
        user.synced_providers = (user.synced_providers or []) + [provider]
        user.save(update_fields=['synced_providers'])
    return JsonResponse({'ok': True})


# ─── Notifications ───
@csrf_exempt
@google_auth
def notification_settings(request):
    u = request.guser
    if request.method == 'GET':
        return JsonResponse({'phone': u.phone, 'whatsapp_optin': u.whatsapp_optin})
    d = _json(request)
    u.phone = d.get('phone') or None
    u.whatsapp_optin = bool(d.get('whatsapp_optin'))
    u.save(update_fields=['phone', 'whatsapp_optin'])
    return JsonResponse({'ok': True})


# ─── Push ───
@csrf_exempt
def vapid_key(request):
    return JsonResponse({'publicKey': settings.VAPID_PUBLIC_KEY})


@csrf_exempt
@google_auth
def push_subscribe(request):
    d = _json(request)
    sub = d.get('subscription')
    if not sub:
        return JsonResponse({'error': 'No subscription'}, status=400)
    PushSubscription.objects.get_or_create(user_id=request.guser.google_id, subscription=json.dumps(sub) if isinstance(sub, dict) else sub)
    return JsonResponse({'ok': True})


@csrf_exempt
@google_auth
def push_unsubscribe(request):
    PushSubscription.objects.filter(user_id=request.guser.google_id).delete()
    return JsonResponse({'ok': True})
